import json
from openpyxl.workbook import Workbook
import datetime
import pytz  # 时间转换

table_header = [
    'sat_name', 'physical_channel', 'proxy_nickname',
    'proxy_location', 'proxy_receive_time', 'server_receive_time', 'raw_data'
]

# 新建一个workbook
wb = Workbook()
ws = wb.active

# 写入表头
for i in range(1, len(table_header)+1):
    ws.cell(row=1, column=i, value=table_header[i-1])

row = 2
timezone = pytz.timezone('Asia/Shanghai')
with open('./jt4g.json') as f:
    while True:
        line = f.readline()
        if not line:
            break
        dic = json.loads(line)
        dic = dic['_source']  # 去除es引擎辅助内容
        dic['proxy_receive_time'] = datetime.datetime.fromtimestamp(dic['proxy_receive_time']/1000, timezone).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        dic['server_receive_time'] = datetime.datetime.fromtimestamp(dic['server_receive_time']/1000, timezone).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        for i in range(1, len(table_header)+1):
            ws.cell(row=row, column=i, value=dic[table_header[i-1]])
        row += 1
        print(row)

wb.save('jt4g.xlsx')
