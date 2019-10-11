import json
from openpyxl.workbook import Workbook
import datetime
import pytz  # 时间转换

table_header = [
    'full_code_name', 'chinese_code_name', 'subsystem', 'server_receive_time',
    'x_text', 'f_text', 'f_double', 'is_suspicious'
]

# 新建一个workbook
wb = Workbook()
ws = wb.active

# 写入表头
for i in range(1, len(table_header) + 1):
    ws.cell(row=1, column=i, value=table_header[i - 1])

row = 2
timezone = pytz.timezone('Asia/Shanghai')
with open('./a.json', encoding='utf-8') as f:
    while True:
        line = f.readline()
        if not line:
            break
        dic = json.loads(line)
        dic = dic['_source']  # 去除es引擎辅助内容
        # 更改时间戳
        dic['server_receive_time'] = datetime.datetime.fromtimestamp(
            dic['server_receive_time'] / 1000,
            timezone).strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        for i in range(1, len(table_header) + 1):
            ws.cell(row=row, column=i, value=dic[table_header[i - 1]])
        row += 1
        if(row % 100 == 0):
            print(row)
print('saving')
wb.save('dslwp_a.xlsx')
